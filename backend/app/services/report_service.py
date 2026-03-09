"""
Report Generation Service for Voice of Care
Handles AI-powered Excel report generation using Claude and openpyxl
"""

import logging
from typing import List, Dict, Any, Optional
from datetime import date, datetime, time, UTC
from sqlalchemy.orm import Session
from sqlalchemy import and_
import uuid
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.visit import Visit
from app.models.beneficiary import Beneficiary
from app.models.worker import Worker
from app.services.bedrock_service import bedrock_service
from app.services.s3 import s3_service
from app.config import settings

logger = logging.getLogger(__name__)


class ReportService:
    """Service for generating Excel reports from visit data"""
    
    def __init__(self):
        """Initialize report service"""
        self.bedrock = bedrock_service
        self.s3 = s3_service
    
    def generate_report(
        self,
        visit_type: str,
        start_date: date,
        end_date: date,
        worker_id: Optional[str],
        db: Session
    ) -> Dict[str, Any]:
        """
        Generate Excel report for visits
        
        Args:
            visit_type: Type of visit (hbnc, anc, pnc)
            start_date: Start date for report period
            end_date: End date for report period
            worker_id: Optional worker ID string (e.g., "AW000001") to filter by specific ASHA worker
            db: Database session
            
        Returns:
            Dictionary with report_id, download_url, and expires_at
            
        Raises:
            ValueError: If no visits found or invalid parameters
            Exception: If report generation fails
        """
        try:
            # Convert worker_id string to database ID if provided
            worker_db_id = None
            if worker_id:
                worker = db.query(Worker).filter(Worker.worker_id == worker_id).first()
                if not worker:
                    raise ValueError(f"Worker with ID {worker_id} not found")
                worker_db_id = worker.id
            
            # Step 1: Query visits from database
            logger.info(
                f"Generating {visit_type} report from {start_date} to {end_date}"
                f"{f' for worker {worker_id}' if worker_id else ''}"
            )
            
            visits = self.query_visits(
                visit_type=visit_type,
                start_date=start_date,
                end_date=end_date,
                worker_id=None,
                db=db
            )
            
            if not visits:
                raise ValueError(
                    f"No {visit_type} visits found for the specified criteria"
                )
            
            logger.info(f"Found {len(visits)} visits for report")
            
            # Step 2: Format visit data for Claude
            visits_data = self._format_visits_for_claude(visits)
            
            # Step 3: Invoke Claude to get structured report data
            report_data = self._invoke_claude_for_report(visits_data)
            
            # Step 4: Build Excel workbook
            excel_bytes = self.build_excel(report_data)
            
            # Step 5: Upload to S3
            report_id = str(uuid.uuid4())
            filename = f"hbnc_report_{start_date}_{end_date}_{report_id}.xlsx"
            s3_key = f"reports/{filename}"
            
            self.upload_report_to_s3(excel_bytes, s3_key)
            
            # Step 6: Generate presigned URL
            download_url = self.generate_presigned_url(s3_key, expiration=900)
            expires_at = datetime.now(UTC).timestamp() + 900
            
            logger.info(f"Report generated successfully: {report_id}")
            
            return {
                "report_id": report_id,
                "download_url": download_url,
                "expires_at": datetime.fromtimestamp(expires_at, UTC).isoformat()
            }
            
        except ValueError:
            # Re-raise validation errors
            raise
        except Exception as e:
            logger.error(f"Failed to generate report: {str(e)}", exc_info=True)
            raise Exception(f"Report generation failed: {str(e)}")
    
    def query_visits(
        self,
        visit_type: str,
        start_date: date,
        end_date: date,
        worker_id: Optional[int],
        db: Session
    ) -> List[Dict[str, Any]]:
        """
        Query visits from database with filters
        
        Args:
            visit_type: Type of visit (hbnc, anc, pnc)
            start_date: Start date for filtering
            end_date: End date for filtering
            worker_id: Optional worker ID filter
            db: Database session
            
        Returns:
            List of visit dictionaries with joined beneficiary and worker data
        """
        # Build query with joins
        query = db.query(
            Visit,
            Beneficiary.first_name.label('beneficiary_first_name'),
            Beneficiary.last_name.label('beneficiary_last_name'),
            Beneficiary.mcts_id,
            Worker.first_name.label('asha_first_name'),
            Worker.last_name.label('asha_last_name')
        ).join(
            Beneficiary, Visit.beneficiary_id == Beneficiary.id
        ).join(
            Worker, Visit.assigned_asha_id == Worker.id
        ).filter(
            and_(
                Visit.visit_type == visit_type,
                Visit.visit_date_time >= datetime.combine(start_date, time.min),
                Visit.visit_date_time <= datetime.combine(end_date, time.max),
                Visit.is_synced == True
            )
        )
        
        # Add worker filter if specified
        if worker_id:
            query = query.filter(Visit.assigned_asha_id == worker_id)
        
        # Order by visit date
        query = query.order_by(Visit.visit_date_time.asc())
        
        # Execute query
        results = query.all()
        
        # Format results
        visits = []
        for row in results:
            visit = row.Visit
            visits.append({
                'visit': visit,
                'beneficiary_first_name': row.beneficiary_first_name,
                'beneficiary_last_name': row.beneficiary_last_name,
                'mcts_id': row.mcts_id,
                'asha_first_name': row.asha_first_name,
                'asha_last_name': row.asha_last_name
            })
        
        return visits
    
    def _format_visits_for_claude(
        self,
        visits: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        Format visit data for Claude prompt
        
        Args:
            visits: List of visit dictionaries from query_visits
            
        Returns:
            List of formatted visit dictionaries for Claude
        """
        formatted_visits = []
        
        for visit_data in visits:
            visit = visit_data['visit']
            
            # Extract answers from visit_data JSON
            answers = []
            if visit.visit_data and 'answers' in visit.visit_data:
                for answer in visit.visit_data['answers']:
                    answer_dict = {
                        'question_id': answer.get('question_id'),
                        'answer': answer.get('answer'),
                    }
                    
                    # Include transcript if available
                    if 'transcript_en' in answer and answer['transcript_en']:
                        answer_dict['transcript'] = answer['transcript_en']
                    elif 'transcript_hi' in answer and answer['transcript_hi']:
                        answer_dict['transcript'] = answer['transcript_hi']
                    
                    answers.append(answer_dict)
            
            formatted_visit = {
                'beneficiary_name': f"{visit_data['beneficiary_first_name']} {visit_data['beneficiary_last_name']}",
                'mcts_id': visit_data['mcts_id'],
                'asha_name': f"{visit_data['asha_first_name']} {visit_data['asha_last_name']}",
                'visit_date': visit.visit_date_time.strftime('%Y-%m-%d'),
                'day_number': visit.day_number,
                'answers': answers
            }
            
            formatted_visits.append(formatted_visit)
        
        return formatted_visits
    
    def _invoke_claude_for_report(
        self,
        visits_data: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """
        Invoke Claude via Bedrock to generate structured report data
        
        Args:
            visits_data: Formatted visit data
            
        Returns:
            Parsed JSON report data from Claude
            
        Raises:
            Exception: If Claude invocation fails after retries
        """
        max_retries = 3
        last_error = None
        
        for attempt in range(max_retries):
            try:
                # Format prompt using bedrock service
                prompt = self.bedrock.format_hbnc_report_prompt(visits_data)
                
                # Invoke Claude
                response = self.bedrock.invoke_claude(
                    prompt=prompt,
                    max_tokens=4096,
                    temperature=0.0
                )
                
                # Parse JSON response
                report_data = self.bedrock.parse_claude_json_response(response['content'])
                
                # Validate that report_data is a dictionary
                if not isinstance(report_data, dict):
                    logger.error(f"Expected dict from parse_claude_json_response, got {type(report_data)}: {report_data}")
                    raise ValueError(f"Invalid report data format: expected dictionary, got {type(report_data).__name__}")
                
                # Validate required fields
                if 'visits' not in report_data:
                    logger.error(f"Missing 'visits' key in report_data: {report_data}")
                    raise ValueError("Invalid report data: missing 'visits' field")
                
                if not isinstance(report_data['visits'], list):
                    logger.error(f"'visits' field is not a list: {type(report_data['visits'])}")
                    raise ValueError("Invalid report data: 'visits' must be a list")
                
                logger.info(f"Successfully parsed report data with {len(report_data['visits'])} visits")
                
                return report_data
                
            except (ValueError, json.JSONDecodeError) as e:
                last_error = e
                logger.warning(f"Attempt {attempt + 1}/{max_retries} failed to parse Claude response: {str(e)}")
                
                if attempt < max_retries - 1:
                    logger.info(f"Retrying Claude invocation (attempt {attempt + 2}/{max_retries})...")
                    continue
                else:
                    logger.error(f"All {max_retries} attempts failed to generate valid report")
                    raise Exception(f"Failed to generate valid report after {max_retries} attempts: {str(last_error)}")
            
            except Exception as e:
                logger.error(f"Unexpected error during Claude invocation: {str(e)}", exc_info=True)
                raise
    
    def build_excel(self, report_data: Dict[str, Any]) -> bytes:
        """
        Build Excel workbook from report data
        
        Args:
            report_data: Structured report data from Claude
            
        Returns:
            Excel file as bytes
        """
        # Validate report_data type
        if not isinstance(report_data, dict):
            logger.error(f"build_excel received invalid type: {type(report_data)}")
            raise TypeError(f"report_data must be a dictionary, got {type(report_data).__name__}")
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "HBNC Report"
        
        # Define headers
        headers = [
            "S.No",
            "Beneficiary Name",
            "MCTS ID",
            "ASHA Worker",
            "Visit Date",
            "Day",
            "Breathing",
            "Feeding",
            "Temperature",
            "Umbilical Cord",
            "Jaundice",
            "Weight (kg)",
            "Remarks"
        ]
        
        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        visits = report_data.get('visits', [])
        
        if not isinstance(visits, list):
            logger.error(f"'visits' field is not a list: {type(visits)}")
            raise TypeError(f"'visits' must be a list, got {type(visits).__name__}")
        for row_num, visit in enumerate(visits, 2):
            ws.cell(row=row_num, column=1, value=visit.get('serial_no', row_num - 1))
            ws.cell(row=row_num, column=2, value=visit.get('beneficiary_name', ''))
            ws.cell(row=row_num, column=3, value=visit.get('mcts_id', ''))
            ws.cell(row=row_num, column=4, value=visit.get('asha_worker', ''))
            ws.cell(row=row_num, column=5, value=visit.get('visit_date', ''))
            ws.cell(row=row_num, column=6, value=visit.get('day_number', ''))
            ws.cell(row=row_num, column=7, value=visit.get('breathing_normal', ''))
            ws.cell(row=row_num, column=8, value=visit.get('feeding_well', ''))
            ws.cell(row=row_num, column=9, value=visit.get('temperature_normal', ''))
            ws.cell(row=row_num, column=10, value=visit.get('umbilical_cord_normal', ''))
            ws.cell(row=row_num, column=11, value=visit.get('jaundice_present', ''))
            ws.cell(row=row_num, column=12, value=visit.get('weight_kg', ''))
            ws.cell(row=row_num, column=13, value=visit.get('remarks', ''))
        
        # Add summary row
        summary_row = len(visits) + 2
        ws.cell(row=summary_row, column=1, value="Total Visits:")
        ws.cell(row=summary_row, column=2, value=len(visits))
        
        # Apply bold to summary row
        summary_font = Font(bold=True)
        ws.cell(row=summary_row, column=1).font = summary_font
        ws.cell(row=summary_row, column=2).font = summary_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()
        
        return excel_bytes
    
    def upload_report_to_s3(self, excel_bytes: bytes, s3_key: str) -> str:
        """
        Upload Excel report to S3
        
        Args:
            excel_bytes: Excel file as bytes
            s3_key: S3 object key
            
        Returns:
            S3 URI
        """
        return self.s3.upload_file(
            file_content=excel_bytes,
            bucket=settings.AWS_S3_BUCKET_REPORTS,
            key=s3_key,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def generate_presigned_url(self, s3_key: str, expiration: int = 900) -> str:
        """
        Generate presigned URL for report download
        
        Args:
            s3_key: S3 object key
            expiration: URL expiration in seconds (default: 900 = 15 minutes)
            
        Returns:
            Presigned URL
        """
        return self.s3.generate_presigned_url(
            bucket=settings.AWS_S3_BUCKET_REPORTS,
            key=s3_key,
            expiration=expiration
        )


# Global report service instance
report_service = ReportService()
